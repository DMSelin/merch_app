from openpyxl.styles import (Alignment, Font, Border, Side)
from datetime import datetime

class FormattingTable:
    def __init__(self, active_sheet):
        self.active_sheet = active_sheet
        self.forma(active_sheet)

    def set_border(self, ws):
        # Все границы для всей таблицы данных
        thin = Side(border_style='thin', color='000000')
        border_table = Border( left = thin, right=thin, top=thin, bottom=thin)
        cells = ws['A4':f'I{ws.max_row}']
        for row in cells:
            for cell in row:
                cell.border = border_table

    def set_font(self, ws):
        # Шрифт для всей таблицы данных
        all_font = Font(name='Calibri', size=14, color='000000')
        header_font = Font(name='Calibri', size=14, color='000000', bold=True)
        all_cells = ws['A1':f'I{ws.max_row}'] # диапозон таблицы без заголовков
        for row in all_cells:
            for cell in row:
                cell.font = all_font
        header_cells = ws['A4' : 'I4'] # диапозон заголовков
        for row in header_cells:
            for cell in row:
                cell.font = header_font
        ws['C2'].font = header_font

    
    def get_num_of_column(self, value):
        print(value)
        d = {'A':0,'B':1,'C':2,'D':3,'E':4,'F':5,'G':6,'H':7,
                'I':8,'J':9,'K':10,'L':11,'M':12,'N':13,'O':14}
        for k, v in d.items():
            if k == value:
                return v

    def column_alignment_center(self, columns):
        for column in columns:
            num_column = self.get_num_of_column(column)
            for row in self.active_sheet[1:self.active_sheet.max_row]:
                column = row[num_column] #Работаем со столбцом в таблице
                column.alignment = Alignment(horizontal='center')

    def column_alignment_left(self, columns):
        for column in columns:
            num_column = self.get_num_of_column(column)
            for row in self.active_sheet[1:self.active_sheet.max_row]:
                column = row[num_column]
                column.alignment = Alignment(horizontal='left')

    def forma(self, active_sheet):
        current_datetime = datetime.now()
        datetoday = f'{current_datetime.day}.{current_datetime.month}.{current_datetime.year}'

        table_name = [
            ('', '', '', '', '', '', '', ''),
            ('', '', f'Отчет {datetoday}', '', '', '', '', ''),
            ('', '', '', '', '', '', '', '')
        ]
        for data in table_name:
            active_sheet.append(data)

        tabel_headers = [('№ п/п', 'Сеть', 'Адрес', 'Наименование', 'Полка', 'Фейсы', 'Цена', 'ДМП (П/Е/С)', 'Примечания')]
        for data in tabel_headers:
            active_sheet.append(data)

        # Ширина столбцов 
        active_sheet.column_dimensions["A"].width = 6
        active_sheet.column_dimensions["B"].width = 18
        active_sheet.column_dimensions["C"].width = 28.0
        active_sheet.column_dimensions["D"].width = 30.0
        active_sheet.column_dimensions["E"].width = 8.0
        active_sheet.column_dimensions["F"].width = 8.0
        active_sheet.column_dimensions["G"].width = 9.0
        active_sheet.column_dimensions["H"].width = 4.0
        active_sheet.column_dimensions["I"].width = 60.0