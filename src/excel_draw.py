import pandas as pd
from datetime import date
from openpyxl import load_workbook, Workbook

import stores_for_report as sfr
import formatting as frmt
import mail 

excel_file = Workbook()
excel_file.create_sheet("Отчет", 0)
active_sheet = excel_file.active

report_sheet = frmt.FormattingTable(active_sheet)

name_stores, name_streets, message_body = mail.mail_parse()
counter = 1
for i in range(0, len(name_stores)):
    if name_stores[i] == "Дикси":
        for data in sfr.diksi(name_streets[i], message_body[i], counter):
            active_sheet.append(data)
    elif name_stores[i] == "Пятёрочка" or name_stores[i] == "Пятерочка" or name_stores[i] == "Пьяница" or name_stores[i] == "Пъятница":
        for data in sfr.pytrchka(name_streets[i], message_body[i], counter):
            active_sheet.append(data)
    elif name_stores[i] == "Перекресток" or name_stores[i] == "Перекрёсток" or name_stores[i] == "Перекрёсток," or name_stores[i] == "Перекресток,": 
        for data in sfr.perek(name_streets[i], message_body[i], counter):
            active_sheet.append(data)
    elif name_stores[i] == "Магнит":
        for data in sfr.magnit(name_streets[i], message_body[i], counter):
            active_sheet.append(data)
    elif name_stores[i] == "КиБ" or name_stores[i] == "Киб" or name_stores[i] == "Кит":
        for data in sfr.kib(name_streets[i], message_body[i], counter):
            active_sheet.append(data)
    elif name_stores[i] == "Верный":
        for data in sfr.verniy(name_streets[i], message_body[i], counter):
            active_sheet.append(data)
    elif name_stores[i] == "Градусы":
        for data in sfr.gradusy(name_streets[i], message_body[i], counter):
            active_sheet.append(data)
    elif name_stores[i] == "Ароматный" or name_stores[i] == "Ам":
        for data in sfr.am(name_streets[i], message_body[i], counter):
            active_sheet.append(data)
    elif name_stores[i] == "Окей":
        for data in sfr.okei(name_streets[i], message_body[i], counter):
            active_sheet.append(data)
    elif name_stores[i] == "Винлаб":
        for data in sfr.vinlab(name_streets[i], message_body[i], counter):
            active_sheet.append(data)
    elif name_stores[i] == "Бристоль":
        for data in sfr.bristol(name_streets[i], message_body[i], counter):
            active_sheet.append(data)
    elif name_stores[i] == "Азбука":
        for data in sfr.Azbyka(name_streets[i], message_body[i], counter):
            active_sheet.append(data)
    elif name_stores[i] == "Мини" or name_stores[i] == "Гипер" or name_stores[i] == "Супер" or name_stores[i] == "Гипермаркет":
        lenta = sfr.lenta(name_stores[i], name_streets[i], message_body[i], counter)
        for data in lenta:
            active_sheet.append(data)
    else:
        print(name_streets[i])
        print("Такого магазина не найдено")
        counter -= 1
    
    counter += 1 

report_sheet.column_alignment_center(['A', 'E', 'F', 'G'])
report_sheet.column_alignment_left(['B', 'C', 'D', 'H', 'I'])

report_sheet.set_border(active_sheet)
report_sheet.set_font(active_sheet)

dt = date.today()
excel_file.save(f'Отчет {dt.day}.{dt.month}.{dt.year} .xlsx')

# Конвертирую Страницу в DataFrame
df = pd.DataFrame(active_sheet.values)